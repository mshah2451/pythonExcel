import pandas as pd
import openpyxl
import numpy as np
import csv
test_str = "Gfg is best for geeks and CS"


print("The original string is : " + str(test_str))
string=[]

wrkbk = openpyxl.load_workbook("Book1.xlsx")
  
sh = wrkbk.active
sub1 = "is"
sub2 = "and"
# iterate through excel and display data
for i in range(1, sh.max_row+1):
    for j in range(1, sh.max_column+1):
        cell_obj = sh.cell(row=i, column=j)
# getting index of substrings
        idx1 = cell_obj.value.find(sub1)
        idx2 = cell_obj.value.find(sub2)
        res = ''
# getting elements in between
        if idx1 != -1:
            for idx in range(idx1 + len(sub1) + 1, idx2):
               res= res + cell_obj.value[idx]
            string.append(res)
print("The extracted string : " + res )
#print(cell_obj.value, end=" ")
with open('Example.csv', 'w', newline = '') as csvfile:
    my_writer = csv.writer(csvfile, delimiter = ',')
    my_writer.writerow(string)

